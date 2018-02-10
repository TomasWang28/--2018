import openpyxl

wb = openpyxl.Workbook()

# 获取活跃的工作表
ws = wb.active

#数据可以直接赋值给单元格
ws['A1'] = 520

#可以整行添加
ws.append([1,2,3])

#Python 类型将自动转换
import datetime
ws['A3'] = datetime.datetime.now()

#保存文件
wb.save("demo.xlsx")

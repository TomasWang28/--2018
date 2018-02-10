import requests
import re
import json
import openpyxl

def open_url(url):
    headers = {"user-agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.98 Safari/537.36"}
    res = requests.get(url, headers=headers)

    return res

def choice_item(page_items):
    print(">>> 可选类别 <<<")
    length = len(page_items)
    for each in range(length):
        print(each+1, page_items[each]['text'])
    choice = int(input("\n请输入类别（数字）："))
    
    return choice

def get_g_page_config(text):
    g_page_config = re.search(r'g_page_config = (.*?);\n', text)
    g_page_config = json.loads(g_page_config.group(1))

    return g_page_config

# 获取排行榜的品牌名称和排行权重
def get_rank(url):
    res = open_url(url)
    g_page_config = get_g_page_config(res.text)
    rank_list = g_page_config['mods']['wbang']['data']['list']
    if rank_list is None:
        return None
    
    brands = []
    for each in rank_list:
        brands.append([each['col2']['text'], each['col4']['percent']])
        
    return brands

def save_to_excel(name, items):
    wb = openpyxl.Workbook()
    
    # 每个小分类保存为一个独立的Excel工作表
    for each_item in items:
        # 获取每个小分类排行榜
        url = "https://top.taobao.com" + each_item['url'][1:] + "&rank=brand&type=hot"
        brands = get_rank(url)
        # 如果该分类下没有数据，则跳过
        if brands is None:
            continue

        # 创建工作表
        ws = wb.create_sheet(title=each_item['text'].replace('/', '&'))
        ws['A1'] = "品牌"
        ws['B1'] = "热门指数"

        # 设置单元格的列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        
        # 将排行榜的数据添加到工作表中
        for each_brand in brands:
            ws.append(each_brand)

        # 制作条形图
        chart1 = openpyxl.chart.BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = each_item['text'].replace('/', '&')
        data = openpyxl.chart.Reference(ws, min_col=2, min_row=1, max_col=2, max_row=(len(brands)+1)) # +1是因为多了一个标题行
        cats = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_col=1, max_row=(len(brands)+1))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws.add_chart(chart1, "D2")

    # 删除默认创建的“Sheet”工作表
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        
    # 保存工作簿
    wb.save(name)

def main():
    url = input("请输入网址：")
    res = open_url(url)
    
    # 获取大分类的名称
    g_page_config = get_g_page_config(res.text)
    page_items = g_page_config['mods']['nav']['data']['common']
          
    # 让用户选择爬取哪一个大分类
    choice = choice_item(page_items)
    
    # 获取该大类别下每个小分类的数据（包括品牌名称和URL等）
    items = page_items[choice-1]['sub']
    
    # 将数据保存为Excel
    name = page_items[choice-1]['text'].replace('/', '&') + '.xlsx'
    save_to_excel(name, items)
        
if __name__ == "__main__":
    main()
